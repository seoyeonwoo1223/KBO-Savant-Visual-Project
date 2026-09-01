"""Build Baseball Savant-style pitcher arsenal profiles from published Excel files."""

from __future__ import annotations

from collections import defaultdict
import json
import math
from pathlib import Path
from statistics import fmean, median

from openpyxl import load_workbook


CM_PER_INCH = 2.54
PITCH_CODES = ("FF", "FT", "SI", "FC", "SL", "ST", "CH", "CU", "FS")
PARK_FACTOR_CODES = ("FF", "SI", "FC", "SL", "CH", "CU", "FS")
PARK_FACTOR_CODE = {"FT": "SI", "ST": "SL"}
PITCH_NAMES = {
    "FF": "포심", "FT": "투심", "SI": "싱커", "FC": "커터", "SL": "슬라이더",
    "ST": "스위퍼", "CH": "체인지업", "CU": "커브", "FS": "포크",
}
PITCH_COLORS = {
    "FF": "#d62f4b", "FT": "#b9415e", "SI": "#f09a22", "FC": "#8d6d61",
    "SL": "#b5b516", "ST": "#3aa8a6", "CH": "#4bb783", "CU": "#76c8c5", "FS": "#7556b8",
}
KOREAN_TO_CODE = {name: code for code, name in PITCH_NAMES.items()}


def _number(value):
    try:
        number = float(value)
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def _season(value) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _stadium(value) -> str:
    name = str(value or "").replace(" ", "")
    aliases = (
        ("고척", "고척"), ("광주", "광주"), ("대구", "대구"),
        ("대전", "대전"), ("한밭", "대전"), ("문학", "문학"), ("인천", "문학"),
        ("사직", "사직"), ("수원", "수원"), ("잠실", "잠실"), ("창원", "창원"),
    )
    return next((canonical for token, canonical in aliases if token in name), name)


def _quantile(values: list[float], probability: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    position = (len(ordered) - 1) * probability
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    return ordered[lower] + (ordered[upper] - ordered[lower]) * (position - lower)


def _summary(values: list[float], digits: int = 1) -> dict | None:
    if not values:
        return None
    return {
        "average": round(fmean(values), digits),
        "low_75": round(_quantile(values, 0.125), digits),
        "high_75": round(_quantile(values, 0.875), digits),
    }


def _load_park_factors(root: Path, season: int) -> dict[tuple[str, str], tuple[float, float]]:
    """Return {(stadium, pitch code): (HB offset cm, IVB offset cm)}.

    The 2022-25 workbooks use a fixed seven-column pitch order.  Some supplied
    header cells contain duplicate labels, so positions are used deliberately;
    the intact 2023 and 2025 files establish the shared order.
    """
    source = root / "data" / "park_adjustments" / f"{season}_VB_Park_Adjustment_v1.0.xlsx"
    if not source.exists():
        raise FileNotFoundError(f"Park adjustment workbook is missing: {source}")
    workbook = load_workbook(source, read_only=True, data_only=True)
    factors: dict[tuple[str, str], list[float | None]] = {}
    try:
        if season >= 2026:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            headers = [str(value or "") for value in next(iterator)]
            for values in iterator:
                row = dict(zip(headers, values))
                code = str(row.get("Pitch") or "").strip()
                hb, ivb = _number(row.get("HB_Offset")), _number(row.get("IVB_Offset"))
                if code in PARK_FACTOR_CODES and hb is not None and ivb is not None:
                    factors[(_stadium(row.get("Stadium")), code)] = [hb, ivb]
        else:
            for metric, index in (("IVB", 1), ("HB", 0)):
                sheet = next(sheet for sheet in workbook.worksheets if metric in sheet.title.upper())
                for values in sheet.iter_rows(min_row=2, values_only=True):
                    stadium = _stadium(values[0])
                    for column, code in enumerate(PARK_FACTOR_CODES, 1):
                        value = _number(values[column] if column < len(values) else None)
                        if value is not None:
                            factors.setdefault((stadium, code), [None, None])[index] = value
    finally:
        workbook.close()
    return {key: (float(value[0]), float(value[1])) for key, value in factors.items() if None not in value}


def _pitch_code(row: dict) -> str:
    code = str(row.get("pitch_type_code") or "").strip().upper()
    if code in PITCH_CODES:
        return code
    return KOREAN_TO_CODE.get(str(row.get("pitch_type_kr") or "").strip(), "")


def _throws(release_x: list[float]) -> str:
    if not release_x:
        return ""
    return "R" if median(release_x) < 0 else "L"


def build_pitch_arsenal(root: Path, season: int, excel_source: Path | None = None) -> tuple[int, int]:
    """Export searchable pitcher profiles with usage, velocity, HB and IVB."""
    source = excel_source or root / "exports" / f"visualbaseball_savant_{season}_latest.xlsx"
    if not source.exists():
        raise FileNotFoundError(f"Pitch Arsenal input workbook is missing: {source}")
    factors = _load_park_factors(root, season)

    workbook = load_workbook(source, read_only=True, data_only=True)
    pitchers: dict[str, dict] = {}
    eligible = 0
    try:
        sheet = workbook["Pitches"]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(iterator, ())]
        for values in iterator:
            row = dict(zip(headers, values))
            if _season(row.get("season")) != season or str(row.get("parse_status") or "") != "ok":
                continue
            code = _pitch_code(row)
            pitcher_name = str(row.get("pitcher_name") or "").strip()
            if not code or not pitcher_name:
                continue
            pitcher_id = str(row.get("pitcher_id") or pitcher_name).strip()
            pitcher = pitchers.setdefault(pitcher_id, {
                "id": pitcher_id, "name": pitcher_name, "release_x": [],
                "pitches": 0, "groups": defaultdict(lambda: {
                    "n": 0, "velocity": [], "hb": [], "ivb": [], "raw_hb": [], "raw_ivb": [],
                    "movement_total": 0, "movement_adjusted": 0,
                }),
            })
            group = pitcher["groups"][code]
            pitcher["pitches"] += 1
            group["n"] += 1
            velocity = _number(row.get("velocity_kmh"))
            release_x = _number(row.get("x0"))
            if velocity is not None:
                group["velocity"].append(velocity)
            if release_x is not None and abs(release_x) >= 0.1:
                pitcher["release_x"].append(release_x)

            raw_hb = _number(row.get("horizontal_movement_cm"))
            raw_ivb = _number(row.get("vertical_movement_cm"))
            if raw_hb is not None and raw_ivb is not None:
                group["movement_total"] += 1
                factor_code = PARK_FACTOR_CODE.get(code, code)
                offset = factors.get((_stadium(row.get("stadium")), factor_code))
                if offset:
                    group["raw_hb"].append(raw_hb / CM_PER_INCH)
                    group["raw_ivb"].append(raw_ivb / CM_PER_INCH)
                    group["hb"].append((raw_hb + offset[0]) / CM_PER_INCH)
                    group["ivb"].append((raw_ivb + offset[1]) / CM_PER_INCH)
                    group["movement_adjusted"] += 1
            eligible += 1
    finally:
        workbook.close()

    output = root / "web" / "data" / "pitch_arsenal" / str(season) / "players"
    output.mkdir(parents=True, exist_ok=True)
    shards = defaultdict(dict)
    player_index = []
    for pitcher_id, pitcher in sorted(pitchers.items(), key=lambda item: item[1]["name"]):
        pitch_types = []
        for code, group in sorted(pitcher["groups"].items(), key=lambda item: item[1]["n"], reverse=True):
            movement_total = group["movement_total"]
            movement_adjusted = group["movement_adjusted"]
            pitch_types.append({
                "code": code,
                "name": PITCH_NAMES[code],
                "color": PITCH_COLORS[code],
                "n": group["n"],
                "usage": round(group["n"] / pitcher["pitches"] * 100, 1),
                "velocity_kmh": _summary(group["velocity"]),
                "horizontal_break_in": _summary(group["hb"]),
                "ivb_in": _summary(group["ivb"]),
                "raw_horizontal_break_in": _summary(group["raw_hb"]),
                "raw_ivb_in": _summary(group["raw_ivb"]),
                "movement_n": movement_adjusted,
                "movement_total_n": movement_total,
                "movement_coverage": round(movement_adjusted / movement_total * 100, 1) if movement_total else 0.0,
                "park_factor_code": PARK_FACTOR_CODE.get(code, code),
            })
        throws = _throws(pitcher["release_x"])
        payload = {
            "schema_version": 1,
            "season": season,
            "source": {
                "workbook": f"exports/{source.name}",
                "park_adjustment": f"data/park_adjustments/{season}_VB_Park_Adjustment_v1.0.xlsx",
            },
            "method": {
                "movement": "park-adjusted HB and IVB; adjusted = measured + stadium/pitch offset",
                "interval": "central 75% (12.5th to 87.5th percentile)",
                "units": {"velocity": "km/h", "movement": "in"},
                "factor_aliases": {"FT": "SI", "ST": "SL"},
            },
            "player": {"id": pitcher_id, "name": pitcher["name"], "throws": throws, "pitches": pitcher["pitches"]},
            "pitch_types": pitch_types,
        }
        shard = pitcher_id[0] if pitcher_id and pitcher_id[0].isdigit() else "other"
        shards[shard][pitcher_id] = payload
        player_index.append({
            "id": pitcher_id, "name": pitcher["name"], "throws": throws,
            "pitches": pitcher["pitches"], "file": f"players/{shard}.json",
        })

    current_files = set()
    for shard, profiles in shards.items():
        filename = f"{shard}.json"
        current_files.add(filename)
        (output / filename).write_text(json.dumps({
            "season": season, "players": profiles,
        }, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    for stale in output.glob("*.json"):
        if stale.name not in current_files:
            stale.unlink()

    season_index = output.parent / "index.json"
    season_index.write_text(json.dumps({
        "season": season, "players": player_index,
    }, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")

    catalog_path = root / "web" / "data" / "pitch_arsenal" / "index.json"
    catalog = {"seasons": []}
    if catalog_path.exists():
        catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    catalog["seasons"] = sorted({*catalog.get("seasons", []), season}, reverse=True)
    catalog_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return eligible, len(pitchers)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=".")
    parser.add_argument("--season", type=int, required=True)
    parser.add_argument("--source")
    args = parser.parse_args()
    rows, players = build_pitch_arsenal(
        Path(args.root).resolve(), args.season,
        Path(args.source).resolve() if args.source else None,
    )
    print(f"exported {rows} pitches for {players} pitcher arsenal profiles")
