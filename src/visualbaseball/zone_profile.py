"""Build compact, filterable pitcher zone profiles from the published Excel file."""

from __future__ import annotations

from collections import defaultdict
import json
import math
from pathlib import Path

from openpyxl import load_workbook


X_MIN = -2.0
X_MAX = 2.0
Z_MIN = 0.0
Z_MAX = 4.5
BUCKET_SIZE = 0.5


def _number(value):
    try:
        number = float(value)
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def _truthy(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value or "").strip().lower() in {"1", "true", "yes", "y"}


def _bucket(value: float, minimum: float, maximum: float) -> int | None:
    if not minimum <= value < maximum:
        return None
    return int((value - minimum) / BUCKET_SIZE)


def _pitch_type(row: dict) -> str:
    return str(row.get("pitch_type_kr") or row.get("pitch_type") or row.get("pitch_type_code") or "기타").strip()


def _batting_result(row: dict) -> tuple[int, int]:
    if not _truthy(row.get("is_pa_terminal")):
        return 0, 0
    result = str(row.get("pa_result") or "").strip()
    if not result or result in {"볼넷", "사구", "고의사"} or "SF" in result or result.endswith("희"):
        return 0, 0
    hit = result.endswith(("안", "이", "삼", "홈")) or "홈런" in result
    return 1, int(hit)


def build_zone_profiles(root: Path, season: int, excel_source: Path | None = None) -> tuple[int, int]:
    """Export compact batter and pitcher JSON profiles plus a shared search index."""
    source = excel_source or root / "exports" / f"visualbaseball_savant_{season}_latest.xlsx"
    if not source.exists():
        raise FileNotFoundError(f"Zone profile input workbook is missing: {source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook["Pitches"]
        iterator = sheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            return 0, 0
        columns = [str(value) if value is not None else "" for value in headers]
        players_by_role: dict[str, dict[str, dict]] = {"batter": {}, "pitcher": {}}
        eligible = 0
        for values in iterator:
            row = {column: value for column, value in zip(columns, values)}
            if row.get("season") != season or str(row.get("parse_status") or "") != "ok":
                continue
            px, pz = _number(row.get("px")), _number(row.get("pz"))
            if px is None or pz is None:
                continue
            x_bin = _bucket(px, X_MIN, X_MAX)
            z_bin = _bucket(pz, Z_MIN, Z_MAX)
            if x_bin is None or z_bin is None:
                continue
            balls = row.get("balls_before")
            strikes = row.get("strikes_before")
            if not isinstance(balls, int) or not isinstance(strikes, int):
                continue
            top, bottom = _number(row.get("sz_top")), _number(row.get("sz_bottom"))
            swing = _truthy(row.get("is_swing"))
            contact = _truthy(row.get("is_contact"))
            in_play = _truthy(row.get("is_in_play"))
            velocity = _number(row.get("velocity_kmh"))
            at_bat, hit = _batting_result(row)
            for role in ("batter", "pitcher"):
                name = str(row.get(f"{role}_name") or "").strip()
                if not name:
                    continue
                player_id = str(row.get(f"{role}_id") or name).strip()
                player = players_by_role[role].setdefault(player_id, {
                    "id": player_id,
                    "name": name,
                    "pitches": 0,
                    "zone_top_sum": 0.0,
                    "zone_bottom_sum": 0.0,
                    "zone_n": 0,
                    "groups": defaultdict(lambda: [0] * 11),
                })
                player["pitches"] += 1
                if top is not None and bottom is not None and top > bottom:
                    player["zone_top_sum"] += top
                    player["zone_bottom_sum"] += bottom
                    player["zone_n"] += 1
                # total, swings, whiffs, contacts, in-play, velo sum/n, zone, pitches, AB, hits
                aggregate = player["groups"][(balls, strikes, _pitch_type(row), x_bin, z_bin)]
                aggregate[0] += 1
                aggregate[1] += int(swing)
                aggregate[2] += int(swing and not contact)
                aggregate[3] += int(contact)
                aggregate[4] += int(in_play)
                if velocity is not None:
                    aggregate[5] += velocity
                    aggregate[6] += 1
                aggregate[7] += int(abs(px) <= 10 / 12 and bottom is not None and top is not None and bottom <= pz <= top)
                aggregate[8] += 1
                aggregate[9] += at_bat
                aggregate[10] += hit
            eligible += 1
    finally:
        workbook.close()

    index_players = {}
    for role, players in players_by_role.items():
        output = root / "web" / "data" / "zones" / str(season) / role
        output.mkdir(parents=True, exist_ok=True)
        current_files = set()
        role_index = []
        shards = defaultdict(dict)
        for player_id, player in sorted(players.items(), key=lambda item: item[1]["name"]):
            shard = player_id[0] if player_id and player_id[0].isdigit() else "other"
            filename = f"{shard}.json"
            current_files.add(filename)
            zone_n = player["zone_n"]
            records = [
                [balls, strikes, pitch_type, x_bin, z_bin, *[round(value, 3) if isinstance(value, float) else value for value in values]]
                for (balls, strikes, pitch_type, x_bin, z_bin), values in player["groups"].items()
            ]
            payload = {
                "schema_version": 1,
                "season": season,
                "role": role,
                "source": f"exports/{source.name}",
                "player": {"id": player_id, "name": player["name"], "file": filename},
                "coordinates": {"x_min": X_MIN, "x_max": X_MAX, "z_min": Z_MIN, "z_max": Z_MAX, "bucket_size": BUCKET_SIZE},
                "strike_zone": {
                    "left": -10 / 12,
                    "right": 10 / 12,
                    "bottom": round(player["zone_bottom_sum"] / zone_n, 3) if zone_n else 1.5,
                    "top": round(player["zone_top_sum"] / zone_n, 3) if zone_n else 3.5,
                },
                "columns": ["balls", "strikes", "pitch_type", "x_bin", "z_bin", "total", "swings", "whiffs", "contacts", "in_play", "velo_sum", "velo_n", "zone", "pitches", "at_bats", "hits"],
                "records": records,
            }
            shards[shard][player_id] = payload
            role_index.append({"id": player_id, "name": player["name"], "file": filename, "pitches": player["pitches"]})
        for shard, shard_players in shards.items():
            (output / f"{shard}.json").write_text(json.dumps({
                "season": season, "role": role, "players": shard_players,
            }, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        for stale in output.glob("*.json"):
            if stale.name not in current_files:
                stale.unlink()
        index_players[role] = role_index

    legacy_output = root / "web" / "data" / "zones" / str(season)
    for legacy_file in legacy_output.glob("*.json"):
        legacy_file.unlink()

    index_path = root / "web" / "data" / "zones" / "index.json"
    catalog = {"seasons": [], "players": {}}
    if index_path.exists():
        catalog = json.loads(index_path.read_text(encoding="utf-8"))
    catalog.setdefault("players", {})[str(season)] = index_players
    catalog["seasons"] = sorted((int(value) for value in catalog["players"]), reverse=True)
    index_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return eligible, sum(len(players) for players in index_players.values())


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=".")
    parser.add_argument("--season", type=int, required=True)
    parser.add_argument("--source")
    args = parser.parse_args()
    rows, players = build_zone_profiles(Path(args.root).resolve(), args.season, Path(args.source).resolve() if args.source else None)
    print(f"exported {rows} pitches for {players} batter/pitcher profiles")
