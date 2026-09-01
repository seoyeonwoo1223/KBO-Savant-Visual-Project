"""Build the current-season leaderboard directly from Visual Baseball PBP."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
import json
import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


TEAM_CODES = {
    "두산": "DOO", "삼성": "SAM", "키움": "KIW", "롯데": "LOT", "한화": "HAN",
    "KIA": "KIA", "KT": "KT", "LG": "LG", "NC": "NC", "SSG": "SSG",
}
TEAM_PARKS = {
    "DOO": "잠실", "LG": "잠실", "KIW": "고척", "SSG": "인천", "KT": "수원",
    "HAN": "대전", "KIA": "광주", "SAM": "대구", "LOT": "사직", "NC": "창원",
}
POSITIONS = {"포": "C", "일": "1B", "이": "2B", "삼": "3B", "유": "SS", "좌": "LF", "중": "CF", "우": "RF", "지": "DH"}
POSITION_RUNS = {"C": 10, "1B": -8, "2B": 3, "3B": -1, "SS": 7, "LF": -5, "CF": 3, "RF": -5, "DH": -15}


def _round(value: float | None, digits: int = 3):
    return None if value is None or not math.isfinite(value) else round(value, digits)


def _ratio(numerator: float, denominator: float, scale: float = 1.0):
    return numerator / denominator * scale if denominator else None


def _is_sacrifice(result: str) -> bool:
    return result.endswith("희") or result.endswith("SF")


def _hit_bases(pa_type: str, result: str) -> int:
    if pa_type == "hr":
        return 4
    if pa_type != "hit":
        return 0
    if result.endswith("삼") and result != "삼안":
        return 3
    if result.endswith("이") and result not in {"이안", "이내안", "이번안"}:
        return 2
    return 1


def _raw_metadata(raw_dir: Path):
    pa_metadata, starters, game_teams = {}, set(), {}
    for path in sorted(raw_dir.glob("*.json")):
        source = json.loads(path.read_text(encoding="utf-8"))
        game = source["gameData"]
        game_id = game["gameId"]
        game_teams[game_id] = {
            "top": TEAM_CODES[game["away"]["team"]],
            "bottom": TEAM_CODES[game["home"]["team"]],
        }
        starters.add((game_id, str(game["away"].get("starter") or "")))
        starters.add((game_id, str(game["home"].get("starter") or "")))
        sequence = 0
        for inning in source.get("pbpData", []):
            for pa in inning.get("pas", []):
                sequence += 1
                pa_metadata[f"{game_id}-{sequence:03d}"] = {
                    "team": TEAM_CODES[inning["team"]],
                    "position": POSITIONS.get(str(pa.get("pos") or "")),
                    "rbi": int(pa.get("rbi") or 0),
                }
    return pa_metadata, starters, game_teams


def _constants(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    yearly, hitters, pitchers = workbook["연도별 상수"], workbook["리그 타자"], workbook["리그 투수"]
    values = {
        "park": {yearly.cell(1, column).value: float(yearly.cell(46, column).value) for column in range(2, 11)},
        "weights": {yearly.cell(1, column).value: float(yearly.cell(46, column).value) for column in range(14, 20)},
        "scale": float(yearly["T46"].value),
        "runs_per_win": float(yearly["U46"].value),
        "fip_constant": float(yearly["V46"].value),
        "league_runs_per_pa": float(yearly["W46"].value),
        "fip_hr": float(yearly["Y46"].value),
        "fip_bb": float(yearly["Z46"].value),
        "fip_k": float(yearly["AA46"].value),
        "league_woba": float(hitters["B46"].value),
        "league_obp": float(hitters["R46"].value),
        "league_slg": float(hitters["S46"].value),
        "league_fip": float(pitchers["Q46"].value),
    }
    workbook.close()
    return values


def _pitch_rates(pitches: pd.DataFrame, player_column: str):
    rates = {}
    for player_id, group in pitches.groupby(player_column, dropna=False):
        swings = group["is_swing"].fillna(False).astype(bool)
        contacts = group["is_contact"].fillna(False).astype(bool)
        outside = group["px"].abs().gt(10 / 12) | group["pz"].lt(group["sz_bottom"]) | group["pz"].gt(group["sz_top"])
        whiffs = swings & ~contacts
        called_strikes = group["pitch_call_code"].eq("T")
        rates[str(player_id)] = {
            "CWS%": _ratio(int((whiffs | called_strikes).sum()), len(group), 100),
            "whiff%": _ratio(int(whiffs.sum()), int(swings.sum()), 100),
            "chase%": _ratio(int((swings & outside).sum()), int(outside.sum()), 100),
        }
    return rates


def _aggregate(root: Path, season: int):
    pitches = pd.read_parquet(root / "data" / "processed" / "pitches.parquet")
    pitches = pitches[(pitches["season"] == season) & pitches["game_date"].notna()].copy()
    terminal = pitches[pitches["is_pa_terminal"]].sort_values(["game_date", "game_id", "inning", "inning_half", "event_seq"])
    metadata, starters, game_teams = _raw_metadata(root / "data" / "raw" / str(season))
    batting, pitching = defaultdict(Counter), defaultdict(Counter)
    batter_games, pitcher_games, pitcher_starts = defaultdict(set), defaultdict(set), defaultdict(set)
    batter_positions, runner_pitcher = defaultdict(Counter), {}

    for row in terminal.itertuples(index=False):
        pa = metadata.get(row.pa_id, {})
        batter_id, pitcher_id = str(row.batter_id), str(row.pitcher_id)
        result, pa_type = str(row.pa_result or ""), str(row.pa_type or "")
        batting[batter_id].update({"PA": 1, "RBI": pa.get("rbi", 0)})
        batting[batter_id]["name"], batting[batter_id]["team"] = row.batter_name, pa.get("team", "")
        batter_games[batter_id].add(row.game_id)
        if pa.get("position"):
            batter_positions[batter_id][pa["position"]] += 1

        bases = _hit_bases(pa_type, result)
        if bases:
            batting[batter_id]["H"] += 1
            if bases == 2: batting[batter_id]["2B"] += 1
            if bases == 3: batting[batter_id]["3B"] += 1
            if bases == 4: batting[batter_id]["HR"] += 1
        if pa_type == "k": batting[batter_id]["SO"] += 1
        if result == "볼넷": batting[batter_id]["BB"] += 1
        if result == "사구": batting[batter_id]["HBP"] += 1
        if result == "고의사": batting[batter_id]["IB"] += 1
        if result.endswith("병"): batting[batter_id]["GDP"] += 1
        if result.endswith("SF"): batting[batter_id]["SF"] += 1
        if pa_type in {"hit", "hr", "k"} or (pa_type == "out" and not _is_sacrifice(result) and result not in {"", "WP"}):
            batting[batter_id]["AB"] += 1

        defense_team = game_teams[row.game_id]["bottom" if row.inning_half == "top" else "top"]
        pitching[pitcher_id]["name"], pitching[pitcher_id]["team"] = row.pitcher_name, defense_team
        pitching[pitcher_id]["TBF"] += 1
        pitcher_games[pitcher_id].add(row.game_id)
        if (row.game_id, str(row.pitcher_name)) in starters:
            pitcher_starts[pitcher_id].add(row.game_id)
        if bases:
            pitching[pitcher_id]["H"] += 1
            if bases == 2: pitching[pitcher_id]["2B"] += 1
            if bases == 3: pitching[pitcher_id]["3B"] += 1
            if bases == 4: pitching[pitcher_id]["HR"] += 1
        if pa_type == "k": pitching[pitcher_id]["SO"] += 1
        if result == "볼넷": pitching[pitcher_id]["BB"] += 1
        if result == "사구": pitching[pitcher_id]["HBP"] += 1
        if result == "고의사": pitching[pitcher_id]["IB"] += 1
        if pa_type not in {"bb", "k", "hr"} and result not in {"", "WP"}:
            pitching[pitcher_id]["BIP"] += 1
        pitching[pitcher_id]["outs"] += max(0, int(row.outs_after or 0) - int(row.outs_before or 0))

        before = [str(value) for value in (row.runner_3b_id_before, row.runner_2b_id_before, row.runner_1b_id_before) if str(value or "").strip()]
        after = {str(value) for value in (row.runner_1b_id_after, row.runner_2b_id_after, row.runner_3b_id_after) if str(value or "").strip()}
        for runner in before:
            runner_pitcher.setdefault((row.game_id, row.inning, row.inning_half, runner), pitcher_id)
        candidates = [runner for runner in before if runner not in after]
        if pa_type == "hr":
            candidates.append(batter_id)
        for runner in candidates[: int(row.runs_on_pitch or 0)]:
            batting[runner]["R"] += 1
            responsible = runner_pitcher.get((row.game_id, row.inning, row.inning_half, runner), pitcher_id)
            pitching[responsible]["RA"] += 1
        for runner in after:
            runner_pitcher.setdefault((row.game_id, row.inning, row.inning_half, runner), pitcher_id)

    return pitches, batting, pitching, batter_games, pitcher_games, pitcher_starts, batter_positions


def _columns(items):
    return [{"key": key, "label": label} for key, label in items]


def _batting_rows(stats_by_player, games, positions, rates, constants, season):
    basic, advanced = [], []
    for player_id, stats in stats_by_player.items():
        pa, ab, hits = stats["PA"], stats["AB"], stats["H"]
        if pa < 200 or not stats["team"]:
            continue
        position = positions[player_id].most_common(1)[0][0] if positions[player_id] else "—"
        team = stats["team"]
        pf = constants["park"][TEAM_PARKS[team]]
        singles = hits - stats["2B"] - stats["3B"] - stats["HR"]
        w = constants["weights"]
        denominator = ab + stats["BB"] + stats["HBP"] - stats["IB"] + stats["SF"]
        woba = _ratio(w["wBB"] * (stats["BB"] - stats["IB"]) + w["wHBP"] * stats["HBP"] + w["w1B"] * singles + w["w2B"] * stats["2B"] + w["w3B"] * stats["3B"] + w["wHR"] * stats["HR"], denominator)
        ba = _ratio(hits, ab)
        obp = _ratio(hits + stats["BB"] + stats["HBP"], ab + stats["BB"] + stats["HBP"] + stats["SF"])
        slg = _ratio(singles + 2 * stats["2B"] + 3 * stats["3B"] + 4 * stats["HR"], ab)
        wraa = ((woba - constants["league_woba"]) / constants["scale"]) * pa / pf
        war = (wraa + POSITION_RUNS.get(position, 0) * pa / 600 + 20 * pa / 600) / constants["runs_per_win"]
        wrc = ((woba - constants["league_woba"]) / constants["scale"] + constants["league_runs_per_pa"]) * pa
        wrc_plus = 100 * ((((woba - constants["league_woba"]) / constants["scale"] + constants["league_runs_per_pa"]) + constants["league_runs_per_pa"] * (1 - pf)) / constants["league_runs_per_pa"])
        player_rates = rates.get(player_id, {})
        row = {
            "Player": stats["name"], "Pos": position, "Team": team, "Year": season,
            "WAR": _round(war, 2), "wRC+": _round(wrc_plus, 1), "G": len(games[player_id]),
            **{key: stats[key] for key in ("PA", "AB", "R", "H", "2B", "3B", "HR", "RBI", "BB", "HBP", "IB", "SO", "GDP", "SF")},
            "BA": _round(ba), "OBP": _round(obp), "SLG": _round(slg), "OPS": _round((obp or 0) + (slg or 0)),
            "wOBA": _round(woba), "PF": pf, "whiff%": _round(player_rates.get("whiff%"), 1), "chase%": _round(player_rates.get("chase%"), 1),
        }
        basic.append(row)
        advanced.append({
            "Player": stats["name"], "Pos": position, "Team": team, "Year": season, "oWAR": row["WAR"], "PA": pa,
            "XBH": stats["2B"] + stats["3B"] + stats["HR"], "wOBA": row["wOBA"], "wRC": _round(wrc, 1), "wRC+": row["wRC+"],
            "OPS+": _round(100 * ((obp or 0) / constants["league_obp"] + (slg or 0) / constants["league_slg"] - 1) / pf, 1),
            "K%": _round(_ratio(stats["SO"], pa, 100), 1), "BB%": _round(_ratio(stats["BB"], pa, 100), 1),
            "BB/K": _round(_ratio(stats["BB"], stats["SO"]), 2), "HR%": _round(_ratio(stats["HR"], pa, 100), 1),
            "BABIP": _round(_ratio(hits - stats["HR"], ab - stats["SO"] - stats["HR"] + stats["SF"])),
            "IsoD": _round((obp or 0) - (ba or 0)), "IsoP": _round((slg or 0) - (ba or 0)),
            "whiff%": row["whiff%"], "chase%": row["chase%"],
        })
    basic.sort(key=lambda row: (-float(row["WAR"] or 0), row["Player"]))
    advanced_by_name = {row["Player"]: row for row in advanced}
    advanced = []
    for rank, row in enumerate(basic, 1):
        row["RK"] = rank
        detail = advanced_by_name[row["Player"]]
        detail["RK"] = rank
        advanced.append(detail)
    return basic, advanced


def _pitching_rows(stats_by_player, games, starts, rates, constants, season):
    basic, advanced = [], []
    for player_id, stats in stats_by_player.items():
        innings = stats["outs"] / 3
        if innings < 50 or not stats["team"]:
            continue
        team = stats["team"]
        pf = constants["park"][TEAM_PARKS[team]] * 0.77 + 0.23
        fip = (constants["fip_hr"] * stats["HR"] + constants["fip_bb"] * (stats["BB"] + stats["HBP"]) + constants["fip_k"] * stats["SO"]) / innings + constants["fip_constant"]
        fip_plus = 100 * (2 - (fip / constants["league_fip"]) * (1 / pf))
        war = ((constants["league_fip"] - fip / pf) * innings / 9 + 20 * innings / 180) / constants["runs_per_win"]
        player_rates = rates.get(player_id, {})
        start_count = len(starts[player_id])
        row = {
            "Player": stats["name"], "Pos": "SP" if start_count >= len(games[player_id]) / 2 else "RP", "Team": team, "Year": season,
            "WAR": _round(war, 2), "G": len(games[player_id]), "GS": start_count, "IP": _round(innings, 1),
            **{key: stats[key] for key in ("RA", "TBF", "H", "2B", "3B", "HR", "BB", "HBP", "IB", "SO")},
            "RA9": _round(_ratio(stats["RA"], innings, 9), 2), "FIP": _round(fip, 2),
            "WHIP": _round(_ratio(stats["H"] + stats["BB"], innings), 2), "FIP+": _round(fip_plus, 1), "PF": _round(pf, 6),
            "BIP": stats["BIP"], "CWS%": _round(player_rates.get("CWS%"), 1), "whiff%": _round(player_rates.get("whiff%"), 1), "chase%": _round(player_rates.get("chase%"), 1),
        }
        basic.append(row)
        advanced.append({key: row[key] for key in ("Player", "Pos", "Team", "Year", "WAR", "G", "IP", "FIP", "FIP+", "BIP", "CWS%", "whiff%", "chase%")})
        advanced[-1].update({"K%": _round(_ratio(stats["SO"], stats["TBF"], 100), 1), "BB%": _round(_ratio(stats["BB"], stats["TBF"], 100), 1)})
    basic.sort(key=lambda row: (-float(row["WAR"] or 0), row["Player"]))
    advanced_by_name = {row["Player"]: row for row in advanced}
    advanced = []
    for rank, row in enumerate(basic, 1):
        row["RK"] = rank
        detail = advanced_by_name[row["Player"]]
        detail["RK"] = rank
        advanced.append(detail)
    return basic, advanced


BAT_COLUMNS = [("RK", "순위"), ("Player", "선수"), ("Pos", "포지션"), ("Team", "팀"), ("Year", "연도"), ("WAR", "WAR*"), ("wRC+", "wRC+"), ("G", "G"), ("PA", "PA"), ("AB", "AB"), ("R", "R"), ("H", "H"), ("2B", "2B"), ("3B", "3B"), ("HR", "HR"), ("RBI", "RBI"), ("BB", "BB"), ("HBP", "HBP"), ("IB", "IB"), ("SO", "SO"), ("GDP", "GDP"), ("SF", "SF"), ("BA", "BA"), ("OBP", "OBP"), ("SLG", "SLG"), ("OPS", "OPS"), ("wOBA", "wOBA"), ("PF", "PF"), ("whiff%", "Whiff%"), ("chase%", "Chase%")]
BAT_ADV_COLUMNS = [("RK", "순위"), ("Player", "선수"), ("Pos", "포지션"), ("Team", "팀"), ("Year", "연도"), ("oWAR", "oWAR*"), ("PA", "PA"), ("XBH", "XBH"), ("wOBA", "wOBA"), ("wRC", "wRC"), ("wRC+", "wRC+"), ("OPS+", "OPS+"), ("K%", "K%"), ("BB%", "BB%"), ("BB/K", "BB/K"), ("HR%", "HR%"), ("BABIP", "BABIP"), ("IsoD", "IsoD"), ("IsoP", "IsoP"), ("whiff%", "Whiff%"), ("chase%", "Chase%")]
PITCH_COLUMNS = [("RK", "순위"), ("Player", "선수"), ("Pos", "보직"), ("Team", "팀"), ("Year", "연도"), ("WAR", "WAR*"), ("G", "G"), ("GS", "GS"), ("IP", "IP"), ("RA", "RA†"), ("TBF", "TBF"), ("H", "H"), ("2B", "2B"), ("3B", "3B"), ("HR", "HR"), ("BB", "BB"), ("HBP", "HBP"), ("IB", "IB"), ("SO", "SO"), ("RA9", "RA9†"), ("FIP", "FIP"), ("WHIP", "WHIP"), ("FIP+", "FIP+"), ("PF", "PF"), ("BIP", "BIP"), ("CWS%", "CWS%"), ("whiff%", "Whiff%"), ("chase%", "Chase%")]
PITCH_ADV_COLUMNS = [("RK", "순위"), ("Player", "선수"), ("Pos", "보직"), ("Team", "팀"), ("Year", "연도"), ("WAR", "WAR*"), ("G", "G"), ("IP", "IP"), ("FIP", "FIP"), ("FIP+", "FIP+"), ("BIP", "BIP"), ("K%", "K%"), ("BB%", "BB%"), ("CWS%", "CWS%"), ("whiff%", "Whiff%"), ("chase%", "Chase%")]


def build_vb_leaderboard(root: Path, season: int = 2026, output: Path | None = None) -> Path:
    pitches, batting, pitching, batter_games, pitcher_games, pitcher_starts, positions = _aggregate(root, season)
    constants = _constants(root / "data" / "leaderboards" / "source" / "constants.xlsx")
    bat_rows, bat_advanced = _batting_rows(batting, batter_games, positions, _pitch_rates(pitches, "batter_id"), constants, season)
    pitch_rows, pitch_advanced = _pitching_rows(pitching, pitcher_games, pitcher_starts, _pitch_rates(pitches, "pitcher_id"), constants, season)
    as_of = str(pitches["game_date"].max())
    pitch_metric_rows = [{key: row[key] for key in ("RK", "Player", "Pos", "Team", "G", "CWS%", "whiff%", "chase%")} for row in pitch_rows]
    payload = {
        "schema_version": 2, "season": season, "as_of": as_of,
        "source": {"name": "Visual Baseball PBP", "games": int(pitches["game_id"].nunique()), "pitches": len(pitches)},
        "notes": [
            "2026 타자·투수 누적값은 Visual Baseball PBP를 직접 재집계했으며 공식 KBO 합계와 일부 차이가 날 수 있습니다.",
            "WAR*는 타자는 수비·주루를 제외하고 포지션 보정만 적용한 추정치, 투수는 FIP 기반 추정치입니다.",
            "† 투수 실점은 승계주자 책임을 PBP 주자 ID로 추적한 값이며 공식 자책점과 다릅니다.",
            "도루·도실·자책점·승패·세이브·홀드와 신뢰 가능한 타구 유형은 원자료 한계로 제외했습니다.",
            "기준일을 검증할 수 없는 2026 OAA 수비 자료는 제외했습니다.",
        ],
        "datasets": [
            {"id": "batting", "title": "기본", "columns": _columns(BAT_COLUMNS), "rows": bat_rows},
            {"id": "batting-advanced", "title": "확장", "columns": _columns(BAT_ADV_COLUMNS), "rows": bat_advanced},
            {"id": "pitching", "title": "기본", "columns": _columns(PITCH_COLUMNS), "rows": pitch_rows},
            {"id": "pitching-advanced", "title": "확장", "columns": _columns(PITCH_ADV_COLUMNS), "rows": pitch_advanced},
            {"id": "pitch-value", "title": "투구 지표", "columns": _columns([("RK", "순위"), ("Player", "선수"), ("Pos", "보직"), ("Team", "팀"), ("G", "G"), ("CWS%", "CWS%"), ("whiff%", "Whiff%"), ("chase%", "Chase%")]), "rows": pitch_metric_rows},
        ],
    }
    output = output or root / "web" / "data" / "leaderboards" / f"{season}.json"
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    catalog_path = output.parent / "index.json"
    catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    catalog["availability"][str(season)] = [dataset["id"] for dataset in payload["datasets"]]
    catalog_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path("."))
    parser.add_argument("--season", type=int, default=2026)
    args = parser.parse_args()
    print(build_vb_leaderboard(args.root, args.season))


if __name__ == "__main__":
    main()
