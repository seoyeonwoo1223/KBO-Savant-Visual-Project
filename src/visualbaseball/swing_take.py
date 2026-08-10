"""RE288-based Swing/Take exports for the Visual Baseball pitch table.

The metric is batter-positive: a positive Decision Run is a decision worth
more than the 2026 KBO average result for the same count and normalized pitch
location.  This module deliberately keeps source-limited pitches out of the
calculation rather than filling their game state by inference.
"""
from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, timezone
from hashlib import sha256
import json
import math
from pathlib import Path
from statistics import mean

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

SEASON = 2026
PLAYER_SLUGS = {"박준순": "park-junsoon", "홍창기": "hong-changki"}
PLATE_HALF_WIDTH_FT = 10 / 12  # ABS strike-zone half width used by Visual Baseball.
GRID_STEP = 0.25
MIN_LOCATION_CELL_PITCHES = 25
REGION_ORDER = ("Heart", "Shadow", "Chase", "Waste")
ACTIONS = ("Swing", "Take")


def _number(value):
    try:
        number = float(value)
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None



def _excel_rows(source: Path, season: int) -> list[dict]:
    """Read the published workbook's authoritative Pitches sheet."""
    if not source.exists():
        raise FileNotFoundError(f"Profile input workbook is missing: {source}")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook["Pitches"]
    except KeyError as error:
        raise ValueError(f"Profile input workbook has no Pitches sheet: {source}") from error
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        return []
    columns = [str(value) if value is not None else "" for value in headers]
    rows = []
    for values in iterator:
        row = {column: value for column, value in zip(columns, values)}
        if row.get("season") == season:
            rows.append(row)
    return rows


def _source_metadata(rows: list[dict], source: Path | None) -> dict:
    updated_at = max((str(row.get("fetched_at") or "") for row in rows), default="")
    return {
        "workbook": source.as_posix() if source else "data/processed/pitches.parquet",
        "sha256": sha256(source.read_bytes()).hexdigest() if source else None,
        "updated_at": updated_at or None,
    }

def _action(row):
    """Classify the source call once, including terminal HBP as a take."""
    code = str(row.get("pitch_call_code") or "").upper()
    if code in {"S", "F", "X"}:
        return "Swing"
    if code in {"B", "T"} or (row.get("is_pa_terminal") and str(row.get("pa_type") or "").lower() == "hbp"):
        return "Take"
    return None


def _relative_location(row):
    px, pz, top, bottom = (_number(row.get(key)) for key in ("px", "pz", "sz_top", "sz_bottom"))
    if None in (px, pz, top, bottom) or top <= bottom:
        return None
    # 0 is plate/zone centre and 1 is the nearest strike-zone edge.
    return px / PLATE_HALF_WIDTH_FT, (pz - (top + bottom) / 2) / ((top - bottom) / 2)


def _region(x, z):
    distance = max(abs(x), abs(z))
    if distance <= 2 / 3:
        return "Heart"
    if distance <= 4 / 3:
        return "Shadow"
    if distance <= 2:
        return "Chase"
    return "Waste"


def _state(row, when):
    outs = row.get(f"outs_{when}")
    bases = row.get(f"base_state_code_{when}")
    balls = row.get(f"balls_{when}")
    strikes = row.get(f"strikes_{when}")
    if not all(isinstance(value, int) for value in (outs, bases, balls, strikes)):
        return None
    if not (0 <= outs <= 3 and 0 <= bases <= 7 and 0 <= balls <= 3 and 0 <= strikes <= 2):
        return None
    return bases, outs, balls, strikes


def _eligible(row):
    if row.get("parse_status") != "ok" or _action(row) is None or _relative_location(row) is None:
        return False
    before, after = _state(row, "before"), _state(row, "after")
    # Three outs is only legal as an after-state; it is an RE of zero.
    return bool(before and after and before[1] < 3)


def _cell(x, z):
    return round(x / GRID_STEP), round(z / GRID_STEP)


def _re288(rows):
    """Estimate each state by realised runs to the end of its half-inning."""
    targets = defaultdict(list)
    by_half = defaultdict(list)
    for row in rows:
        by_half[(row["game_id"], row["inning"], row["inning_half"])].append(row)
    for half_rows in by_half.values():
        future_runs = 0.0
        for row in sorted(half_rows, key=lambda item: item["event_seq"], reverse=True):
            future_runs += float(row.get("runs_on_pitch") or 0)
            row["_runs_to_end"] = future_runs
            targets[_state(row, "before")].append(future_runs)
    return {state: mean(values) for state, values in targets.items()}, {state: len(values) for state, values in targets.items()}


def _location_baseline(rows):
    """Count × relative-location-cell mean Raw RV; cells under 25 use neighbours."""
    cells = defaultdict(list)
    for row in rows:
        x, z = row["_relative_location"]
        balls, strikes = row["balls_before"], row["strikes_before"]
        cells[(balls, strikes, *_cell(x, z))].append(row["raw_run_value"])
    result = {}
    for key, values in cells.items():
        balls, strikes, cx, cz = key
        pool = list(values)
        radius = 1
        while len(pool) < MIN_LOCATION_CELL_PITCHES and radius <= 4:
            pool = [value for (b, s, x, z), values in cells.items() if (b, s) == (balls, strikes) and max(abs(x - cx), abs(z - cz)) <= radius for value in values]
            radius += 1
        result[key] = mean(pool)
    return result


def _summary(rows, name, slug, season, re_counts, excluded, source_metadata):
    groups = {(region, action): [] for region in REGION_ORDER for action in ACTIONS}
    for row in rows:
        groups[(row["attack_region"], row["decision_type"])].append(row)
    def aggregate(items):
        pitches = len(items); decision = sum(item["decision_run"] for item in items)
        return {"pitches": pitches, "decision_run": round(decision, 4), "decision_run_per_100": round(100 * decision / pitches, 4) if pitches else None}
    total = aggregate(rows)
    by_region = {}
    for region in REGION_ORDER:
        combined = groups[(region, "Swing")] + groups[(region, "Take")]
        entry = aggregate(combined)
        entry.update({"share_pct": round(100 * len(combined) / len(rows), 3) if rows else 0, "swing": aggregate(groups[(region, "Swing")]), "take": aggregate(groups[(region, "Take")])})
        by_region[region] = entry
    zone_grid = {}
    for row in rows:
        x, z = row["_relative_location"]
        if abs(x) <= 1 and abs(z) <= 1:
            key = f"{min(2, max(0, int((x + 1) * 1.5)))}-{min(2, max(0, int((z + 1) * 1.5)))}"
            zone_grid.setdefault(key, []).append(row)
    return {"schema_version": 2, "metric": "RE288 location-and-count-neutral Decision Run", "season": season, "source": source_metadata, "player": {"name": name, "slug": slug}, "sample": {"eligible_pitches": len(rows), "excluded_pitches": excluded, "minimum_pitches": 100, "meets_minimum": len(rows) >= 100}, "coordinate_contract": {"source": "Visual Baseball ABS px/pz (feet)", "x_center": 0, "x_zone_edge_ft": PLATE_HALF_WIDTH_FT, "z_center": "(sz_top + sz_bottom) / 2", "relative_distance": "max(abs(x_relative), abs(z_relative))", "regions": {"Heart": "0–66.7%", "Shadow": "66.7–133.3%", "Chase": "133.3–200%", "Waste": ">200%"}}, "baseline": {"count": "balls_before × strikes_before", "location": f"normalized {GRID_STEP:.2f} × {GRID_STEP:.2f} cells; sparse cells expand to Chebyshev radius 1.00", "minimum_cell_pitches": MIN_LOCATION_CELL_PITCHES}, "re288": {"observed_states": len(re_counts), "state_counts": {"-".join(map(str, state)): count for state, count in sorted(re_counts.items())}}, "overall": total, "regions": by_region, "zone_grid": {key: aggregate(value) for key, value in zone_grid.items()}}


def build_swing_take(root: Path, season: int = SEASON, excel_source: Path | None = None) -> tuple[int, int]:
    """Build profiles from the published Excel workbook, never a parallel raw input."""
    source = excel_source or root / "data" / "processed" / "pitches.parquet"
    rows = _excel_rows(source, season) if excel_source else [
        dict(row) for row in pq.read_table(source).to_pylist() if row.get("season") == season
    ]
    source_metadata = _source_metadata(rows, excel_source)
    excluded = Counter(); valid = []
    for row in rows:
        if not _eligible(row):
            excluded["source_or_required_state_missing"] += 1
            continue
        row["_relative_location"] = _relative_location(row)
        valid.append(row)
    re288, counts = _re288(valid)
    processed = root / "data" / "processed"
    processed.mkdir(parents=True, exist_ok=True)
    (processed / "re288.json").write_text(json.dumps({
        "season": season,
        "states": [{"base_state_code": state[0], "outs": state[1], "balls": state[2], "strikes": state[3], "run_expectancy": round(value, 6), "pitches": counts[state]} for state, value in sorted(re288.items())],
    }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    valued = []
    for row in valid:
        before, after = _state(row, "before"), _state(row, "after")
        if before not in re288 or (after[1] < 3 and after not in re288):
            excluded["unobserved_re288_state"] += 1; continue
        row["raw_run_value"] = float(row.get("runs_on_pitch") or 0) + (0 if after[1] == 3 else re288[after]) - re288[before]
        valued.append(row)
    baseline = _location_baseline(valued)
    output_rows = []
    for row in valued:
        x, z = row["_relative_location"]; key = (row["balls_before"], row["strikes_before"], *_cell(x, z))
        row.update({"x_relative": x, "z_relative": z, "attack_region": _region(x, z), "decision_type": _action(row), "decision_run": row["raw_run_value"] - baseline[key], "location_count_baseline": baseline[key]})
        output_rows.append({key: value for key, value in row.items() if not key.startswith("_")})
    pq.write_table(pa.Table.from_pylist(output_rows), processed / "decision_pitches.parquet")
    output = root / "web" / "data" / "profiles"; output.mkdir(parents=True, exist_ok=True)
    for name, slug in PLAYER_SLUGS.items():
        profile_rows = [row for row in valued if row.get("batter_name") == name]
        (output / f"{slug}.json").write_text(json.dumps(_summary(profile_rows, name, slug, season, counts, dict(excluded), source_metadata), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return len(output_rows), sum(1 for row in valued if row.get("batter_name") in PLAYER_SLUGS)
