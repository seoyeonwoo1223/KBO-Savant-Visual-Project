from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


STADIUM_ALIASES = {"문학": "인천"}


def calculate_run_park_factors(workbook_path: Path, season: int) -> dict[str, float]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Games"]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        index = {name: position for position, name in enumerate(headers)}
        games = [
            {name: values[position] for name, position in index.items()}
            for values in rows
            if values[index["season"]] == season and values[index["is_final"]]
        ]
    finally:
        workbook.close()

    home_stadium_counts: dict[str, Counter] = defaultdict(Counter)
    for game in games:
        home_stadium_counts[game["home_team"]][game["stadium"]] += 1
    primary_stadium = {
        team: counts.most_common(1)[0][0]
        for team, counts in home_stadium_counts.items()
    }
    stadium_teams: dict[str, list[str]] = defaultdict(list)
    for team, stadium in primary_stadium.items():
        stadium_teams[stadium].append(team)

    factors = {}
    for stadium, teams in stadium_teams.items():
        home_games = [
            game for game in games
            if game["stadium"] == stadium and game["home_team"] in teams
        ]
        away_games = [
            game for game in games
            if game["away_team"] in teams and game["stadium"] != stadium
        ]
        if not home_games or not away_games:
            continue
        home_runs_per_game = sum(
            game["home_score"] + game["away_score"] for game in home_games
        ) / len(home_games)
        away_runs_per_game = sum(
            game["home_score"] + game["away_score"] for game in away_games
        ) / len(away_games)
        factors[STADIUM_ALIASES.get(stadium, stadium)] = round(
            home_runs_per_game / away_runs_per_game, 6
        )
    return factors
