"""Build compact historical batter-handedness metadata from Visual Baseball PBP."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
from pathlib import Path
from threading import local

from openpyxl import load_workbook

from visualbaseball.http_client import VisualBaseballClient


def _games(root: Path, season: int) -> list[str]:
    workbook = load_workbook(
        root / "exports" / f"visualbaseball_savant_{season}_latest.xlsx",
        read_only=True,
        data_only=True,
    )
    try:
        rows = workbook["Games"].iter_rows(values_only=True)
        headers = list(next(rows))
        game_id = headers.index("game_id")
        return sorted({str(row[game_id]) for row in rows if row[game_id]})
    finally:
        workbook.close()


def _raw_path(root: Path, season: int, game_id: str) -> Path | None:
    candidates = (
        root / "data" / "raw" / str(season) / f"{game_id}.json",
        root / "seasons" / str(season) / "data" / "raw" / str(season) / f"{game_id}.json",
    )
    return next((path for path in candidates if path.exists()), None)


def _observe(payload: dict, counts: dict[str, Counter], names: dict[str, str]) -> None:
    for half in payload.get("pbpData", []):
        for pa in half.get("pas") or []:
            batter_id = str(pa.get("batterId") or "").strip()
            if not batter_id:
                continue
            names[batter_id] = str(pa.get("batter") or names.get(batter_id, ""))
            for pitch in pa.get("pitches") or []:
                stance = str(pitch.get("stance") or "").strip().upper()
                if stance in {"L", "R"}:
                    counts[batter_id][stance] += 1


def _bats(observed: Counter) -> str:
    """Treat a tiny contradictory tail as source noise, not switch-hitting."""
    left, right = observed["L"], observed["R"]
    minority, total = min(left, right), left + right
    if left and right and minority >= 20 and minority / total >= 0.05:
        return "S"
    return "L" if left >= right else "R"


def build(root: Path, seasons: list[int], workers: int) -> Path:
    output = {"schema_version": 1, "source": "Visual Baseball PBP pitch.stance", "seasons": {}}
    for season in seasons:
        game_ids = _games(root, season)
        counts: dict[str, Counter] = defaultdict(Counter)
        names: dict[str, str] = {}
        missing: list[str] = []
        for game_id in game_ids:
            path = _raw_path(root, season, game_id)
            if path:
                _observe(json.loads(path.read_text(encoding="utf-8-sig")), counts, names)
            else:
                missing.append(game_id)

        state = local()

        def fetch(game_id: str) -> dict:
            client = getattr(state, "client", None)
            if client is None:
                client = VisualBaseballClient()
                state.client = client
            return client.get_json(f"/api/game/pbp?id={game_id}", f"/game/{game_id}/pbp")

        with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
            futures = {executor.submit(fetch, game_id): game_id for game_id in missing}
            for completed, future in enumerate(as_completed(futures), 1):
                _observe(future.result(), counts, names)
                if completed % 50 == 0 or completed == len(futures):
                    print(f"{season}: fetched {completed}/{len(futures)} games", flush=True)

        players = {}
        for batter_id, observed in sorted(counts.items()):
            bats = _bats(observed)
            players[batter_id] = {
                "name": names.get(batter_id, ""),
                "bats": bats,
                "observed": {"L": observed["L"], "R": observed["R"]},
            }
        output["seasons"][str(season)] = {"games": len(game_ids), "players": players}
        print(f"{season}: {len(players)} batters", flush=True)

    target = root / "data" / "batter_handedness.json"
    target.write_text(json.dumps(output, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return target


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=".")
    parser.add_argument("--seasons", nargs="+", type=int, default=[2022, 2023, 2024, 2025])
    parser.add_argument("--workers", type=int, default=8)
    args = parser.parse_args()
    print(build(Path(args.root).resolve(), args.seasons, args.workers))


if __name__ == "__main__":
    main()
