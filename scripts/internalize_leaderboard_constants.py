"""Embed leaderboard constants in a season workbook and remove external links."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


CONSTANT_SHEETS = ("연도별 상수", "리그 타자", "리그 투수")


def copy_sheet(source, target) -> None:
    for row in source.iter_rows():
        for source_cell in row:
            target_cell = target[source_cell.coordinate]
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)

    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key] = copy(dimension)
    for key, dimension in source.row_dimensions.items():
        target.row_dimensions[key] = copy(dimension)
    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))

    target.sheet_format = copy(source.sheet_format)
    target.sheet_properties = copy(source.sheet_properties)
    target.freeze_panes = source.freeze_panes
    target.sheet_state = "hidden"


def trim_phantom_cells(workbook) -> None:
    """Discard empty cells stored far outside each sheet's real data range."""
    for sheet in workbook.worksheets:
        valued = [key for key, cell in sheet._cells.items() if cell.value is not None]
        if not valued:
            sheet._cells.clear()
            continue

        max_row = max(row for row, _ in valued)
        max_column = max(column for _, column in valued)
        for row, column in list(sheet._cells):
            if row > max_row or column > max_column:
                del sheet._cells[(row, column)]


def align_detail_sheets(workbook) -> None:
    """Drop trailing formula-only rows that have no matching player row."""
    for basic_name, detail_name in (("기본-타자", "확장-타자"), ("기본-투수", "확장-투수")):
        basic = workbook[basic_name]
        detail = workbook[detail_name]
        basic_rows = max(cell.row for cell in basic._cells.values() if cell.value is not None)
        if detail.max_row > basic_rows:
            detail.delete_rows(basic_rows + 1, detail.max_row - basic_rows)


def internalize(workbook_path: Path, constants_path: Path, output_path: Path) -> None:
    workbook = load_workbook(workbook_path, data_only=False, keep_links=True)
    constants = load_workbook(constants_path, data_only=False)

    for sheet_name in CONSTANT_SHEETS:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        copy_sheet(constants[sheet_name], workbook.create_sheet(sheet_name))

    replacements = 0
    for sheet in workbook.worksheets:
        for cell in sheet._cells.values():
            if isinstance(cell.value, str) and cell.value.startswith("=") and "[1]" in cell.value:
                cell.value = cell.value.replace("[1]", "")
                replacements += 1

    align_detail_sheets(workbook)
    trim_phantom_cells(workbook)
    workbook._external_links = []
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"internalized {replacements} formulas into {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("constants", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    internalize(args.workbook, args.constants, args.output)


if __name__ == "__main__":
    main()
